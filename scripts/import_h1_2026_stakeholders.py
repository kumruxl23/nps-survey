"""Import H1 2026 targeted-stakeholder nominations from per-org workbooks.

Each WHS team maintains its own targeted-stakeholder workbook. They share the
same overall shape (an "H1 2026" sheet + a "Stakeholder List" sheet), but
column names differ slightly between teams. This script handles all three:

  - whs_cpt_in   :  Sandeep Directs - WHS CPT IN Targeted Stakeholder List ...
  - whs_cpt_na   :  WHS CPT -NA Targeted Stakeholder List for 2025 NPS Survey
  - fec          :  FEC - NPS Targeted Stakeholder List

Org-specific column names are captured in ORG_CONFIGS below. Add a new entry
to onboard another org.

For each "Yes"-flagged stakeholder on the H1 2026 sheet, this writes a
``Nomination`` row. Leader is read from the Stakeholder List sheet (which
has the org's leader-mapping column) where available, falling back to the
POC name from H1 2026.

Usage (run on the EC2):

    /usr/bin/python3.11 scripts/import_h1_2026_stakeholders.py \\
        --workbook /tmp/cpt_in.xlsx --org whs_cpt_in --dry-run

    /usr/bin/python3.11 scripts/import_h1_2026_stakeholders.py \\
        --workbook /tmp/fec.xlsx --org fec --dry-run

    /usr/bin/python3.11 scripts/import_h1_2026_stakeholders.py \\
        --workbook /tmp/cpt_na.xlsx --org whs_cpt_na --dry-run

Drop --dry-run to actually write.

The script writes ONLY to NpsNominations. It does NOT touch responses,
cycles, or org configs.
"""

from __future__ import annotations

import argparse
import logging
import os
import sys
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
os.environ.setdefault("AWS_REGION", "ap-south-1")
os.environ.setdefault("AWS_DEFAULT_REGION", "ap-south-1")

from app.db import nps_cycle_repo, nps_nomination_repo  # noqa: E402
from app.db.models import Nomination  # noqa: E402
from app.services.nomination_keys import encode_for_leader  # noqa: E402

logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
logger = logging.getLogger("import_stakeholders")

DEFAULT_CYCLE_ID = "h1-2026"

# Headers shared by all three orgs (same exact text).
INCLUSION_COL_TEXT = "Should this stakeholder be included in\nH1 2026 survey?\n(Targeted list)"
RESPONDED_COL_TEXT = "Did stakeholder respond to\nH1 NPS 2026?"

# Per-org column conventions. The 'h1_sheet' map describes the *H1 2026* sheet's
# column headers; the 'stakeholder_list' map describes the *Stakeholder List*
# sheet's columns by position (since some have duplicate or trailing-whitespace
# header text that's awkward to match by string).
ORG_CONFIGS: dict[str, dict] = {
    "whs_cpt_in": {
        "h1_sheet_name": "H1 2026",
        "h1_sheet_cols": {
            "poc_name": "PoC",
            "stakeholder_name": "Stakeholder",
            "stakeholder_alias": "Stakeholder\nAlias",  # newline in header
        },
        "stakeholder_list_sheet_name": "Stakeholder List",
        # In WHS CPT IN's workbook there is no separate "Leader" column;
        # POC IS the leader. Position-based read for safety.
        # Cols (0-indexed): 0=PoC, 1=PoC Alias, 2/3=Stakeholder name/alias mix, 4=Email
        "stakeholder_list_positions": {
            "leader_pos": 0,
            "stakeholder_alias_pos": 3,
            "email_pos": 4,
        },
    },
    "fec": {
        "h1_sheet_name": "H1 2026",
        "h1_sheet_cols": {
            "poc_name": "WHS FEC POC Name",
            "stakeholder_name": "Stakeholder Name",
            "stakeholder_alias": "Stakeholder Login/Alias",
        },
        "stakeholder_list_sheet_name": "Stakeholder List",
        # FEC Stakeholder List cols: 0=WHS FEC Leader, 1=WHS FEC POC Name,
        # 2=Stakeholder Name, 3=Stakeholder Login/Alias, 4="Email " (trailing space), 5=Responded
        "stakeholder_list_positions": {
            "leader_pos": 0,
            "stakeholder_alias_pos": 3,
            "email_pos": 4,
        },
    },
    "whs_cpt_na": {
        "h1_sheet_name": "H1 2026",
        "h1_sheet_cols": {
            "poc_name": "WHS CP PoC Name",
            "stakeholder_name": "Stakeholder Name",
            "stakeholder_alias": "Stakeholder Login/Alias",
        },
        "stakeholder_list_sheet_name": "Stakeholder List",
        # CPT NA Stakeholder List cols: 0=CP NA Team, 1=WHS CP PoC Name,
        # 2=Stakeholder Name, 3=Stakeholder Login/Alias, 4=Email, 5=Responded
        "stakeholder_list_positions": {
            "leader_pos": 0,
            "stakeholder_alias_pos": 3,
            "email_pos": 4,
        },
    },
}


def _norm_leader_name(name: str) -> str:
    """Collapse spelling variations (e.g. 'NIdhi Bhagat' -> 'Nidhi Bhagat')."""
    if not name:
        return ""
    cleaned = " ".join(str(name).split())
    return cleaned.title()


def _row_to_dict(headers: list[str], row_values: list) -> dict[str, object]:
    return {h: v for h, v in zip(headers, row_values) if h is not None}


def _find_header_row(ws, max_check: int = 10) -> int:
    """Return 1-indexed row number of the first row with >=3 non-empty cells."""
    for r in range(1, min(ws.max_row, max_check) + 1):
        nonempty = sum(1 for c in ws[r] if c.value not in (None, ""))
        if nonempty >= 3:
            return r
    raise ValueError("No header row found in first 10 rows")


def _parse_h1_2026(workbook_path: Path, cfg: dict) -> list[dict]:
    """Return rows from the H1 2026 sheet that are flagged Yes for inclusion."""
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    sheet_name = cfg["h1_sheet_name"]
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Workbook missing sheet '{sheet_name}'. Found: {wb.sheetnames}")
    ws = wb[sheet_name]
    header_row = _find_header_row(ws)
    headers = [c.value for c in ws[header_row]]

    poc_col = cfg["h1_sheet_cols"]["poc_name"]
    name_col = cfg["h1_sheet_cols"]["stakeholder_name"]
    alias_col = cfg["h1_sheet_cols"]["stakeholder_alias"]

    yes_rows: list[dict] = []
    for r in range(header_row + 1, ws.max_row + 1):
        values = [c.value for c in ws[r]]
        if not any(v not in (None, "") for v in values):
            continue
        row = _row_to_dict(headers, values)

        flag = (row.get(INCLUSION_COL_TEXT) or "")
        if str(flag).strip().lower() != "yes":
            continue

        yes_rows.append({
            "poc": _norm_leader_name(str(row.get(poc_col) or "")),
            "stakeholder": str(row.get(name_col) or "").strip(),
            "stakeholder_alias": str(row.get(alias_col) or "").strip().lower(),
            "responded": str(row.get(RESPONDED_COL_TEXT) or "").strip().lower(),
        })

    return yes_rows


def _parse_stakeholder_list(workbook_path: Path, cfg: dict) -> dict[str, dict]:
    """Return alias -> {email, leader} from the Stakeholder List sheet.

    Reads by column *position* rather than by header text — the headers vary
    and sometimes contain duplicates / trailing whitespace.
    """
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    sheet_name = cfg["stakeholder_list_sheet_name"]
    if sheet_name not in wb.sheetnames:
        logger.warning("Workbook missing '%s' sheet; emails will be constructed from aliases.",
                       sheet_name)
        return {}

    ws = wb[sheet_name]
    pos = cfg["stakeholder_list_positions"]
    leader_pos = pos["leader_pos"]
    alias_pos = pos["stakeholder_alias_pos"]
    email_pos = pos["email_pos"]

    header_row = _find_header_row(ws)
    by_alias: dict[str, dict] = {}

    for r in range(header_row + 1, ws.max_row + 1):
        values = [c.value for c in ws[r]]
        if not any(v not in (None, "") for v in values):
            continue
        leader = _norm_leader_name(str(values[leader_pos]) if leader_pos < len(values) and values[leader_pos] else "")
        stakeholder_alias = (str(values[alias_pos]) if alias_pos < len(values) and values[alias_pos] else "").strip().lower()
        email_raw = values[email_pos] if email_pos < len(values) else None
        if not stakeholder_alias:
            continue
        email = (str(email_raw) if email_raw else "").strip().lower()
        if not email:
            email = f"{stakeholder_alias}@amazon.com"

        by_alias[stakeholder_alias] = {
            "email": email,
            "leader": leader,
        }

    return by_alias


# Helpers for multi-leader encoding -----------------------------------------
# When the same stakeholder is nominated by multiple leaders, we need
# multiple Nomination rows. The DynamoDB sort key is `email`, so for the
# 2nd, 3rd, ... leader of a shared stakeholder we suffix the email with
# `#<leader-name>` to keep keys unique. Always: the FIRST leader gets
# the plain email; later ones get the suffix. Reminder dispatch strips
# suffixes before sending so the human only gets one email.

NOM_SUFFIX_SEP = "#"


def encode_nomination_email(email: str, leader: str, taken_keys: set[str]) -> str:
    """Pick a unique sort-key for a nomination row (deprecated wrapper).

    Use ``app.services.nomination_keys.encode_for_leader`` going forward.
    """
    return encode_for_leader(email, leader, taken_keys)


def decode_nomination_email(suffixed: str) -> str:
    """Strip the ``#leader`` suffix to recover the human-deliverable email."""
    if not suffixed:
        return suffixed
    return suffixed.split(NOM_SUFFIX_SEP, 1)[0]


def import_nominations(
    workbook_path: Path,
    org_id: str,
    cycle_id: str,
    dry_run: bool,
) -> dict[str, int]:
    if org_id not in ORG_CONFIGS:
        raise ValueError(f"Unknown org '{org_id}'. Known: {list(ORG_CONFIGS.keys())}")
    cfg = ORG_CONFIGS[org_id]

    stats = {
        "yes_rows": 0,
        "matched_in_stakeholder_list": 0,
        "fallback_email_constructed": 0,
        "fallback_leader_from_poc": 0,
        "nominations_written": 0,
        "shared_with_suffix": 0,
    }

    yes_rows = _parse_h1_2026(workbook_path, cfg)
    stakeholder_map = _parse_stakeholder_list(workbook_path, cfg)
    stats["yes_rows"] = len(yes_rows)
    logger.info("[%s] Parsed %d 'Yes' rows from %s", org_id, len(yes_rows), cfg["h1_sheet_name"])
    logger.info("[%s] Loaded %d entries from %s",
                org_id, len(stakeholder_map), cfg["stakeholder_list_sheet_name"])

    existing_cycle = nps_cycle_repo.get_cycle(org_id, cycle_id)
    if not existing_cycle:
        logger.warning("[%s] cycle %s does not exist. Run backfill_from_asana --backfill "
                       "first to create the active cycle row, then re-run this script.",
                       org_id, cycle_id)
        return stats

    # Track which (email, leader) pairs we've already written so we don't
    # double-import the *same* relationship if the workbook duplicates a row
    # by accident, while still allowing the same email under DIFFERENT leaders.
    taken_keys: set[str] = set()
    seen_pairs: set[tuple[str, str]] = set()

    for row in yes_rows:
        alias = row["stakeholder_alias"]
        if not alias:
            continue

        sh = stakeholder_map.get(alias)
        if sh and sh.get("email"):
            email = sh["email"].strip().lower()
            stats["matched_in_stakeholder_list"] += 1
        else:
            email = f"{alias}@amazon.com"
            stats["fallback_email_constructed"] += 1
            logger.debug("[%s] no stakeholder-list email for alias '%s', constructed '%s'",
                         org_id, alias, email)

        # Leader source of truth = H1 2026 sheet's POC column. The
        # Stakeholder List sheet only carries one leader per alias, which
        # silently collapses shared stakeholders (e.g. Karthik Rao under
        # both Abhas Rao AND Indrajeet Roy in cpt_in). The H1 sheet
        # encodes the multi-leader truth via duplicate rows with
        # different POC values.
        leader = row["poc"] or (sh.get("leader") if sh else "")
        if not leader and sh and sh.get("leader"):
            stats["fallback_leader_from_poc"] += 1

        pair = (email, leader)
        if pair in seen_pairs:
            logger.debug("[%s] duplicate (email, leader) pair %s, skipping", org_id, pair)
            continue
        seen_pairs.add(pair)

        # Multi-leader encoding: same email under a *different* leader gets
        # the `#leader` suffix to keep DynamoDB keys unique.
        was_already_taken = email in taken_keys
        nomination_email = encode_for_leader(email, leader, taken_keys)
        if was_already_taken:
            stats["shared_with_suffix"] += 1

        nomination = Nomination(
            org_id=org_id,
            cycle_id=cycle_id,
            email=nomination_email,
            name=row["stakeholder"],
            leader=leader,
            responded=(row["responded"] == "yes"),
            responded_at=("" if row["responded"] != "yes"
                          else datetime.now(timezone.utc).isoformat()),
        )

        if dry_run:
            logger.debug("[%s] DRY RUN nomination: leader='%s' email=%s name='%s' responded=%s",
                         org_id, nomination.leader, nomination.email, nomination.name, nomination.responded)
        else:
            nps_nomination_repo.put_nomination(nomination)

        stats["nominations_written"] += 1

    return stats


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", type=Path, required=True,
                        help="Path to the team's stakeholder workbook (.xlsx).")
    parser.add_argument("--org", required=True, choices=sorted(ORG_CONFIGS.keys()),
                        help="org_id to import nominations under.")
    parser.add_argument("--cycle-id", default=DEFAULT_CYCLE_ID,
                        help="cycle_id to attach nominations to (default: h1-2026).")
    parser.add_argument("--dry-run", action="store_true",
                        help="Read-only — print stats, write nothing.")
    args = parser.parse_args()

    if not args.workbook.is_file():
        sys.exit(f"Workbook not found: {args.workbook}")

    logger.info("=== Importing H1 2026 nominations (org=%s, cycle=%s, dry_run=%s) ===",
                args.org, args.cycle_id, args.dry_run)
    stats = import_nominations(args.workbook, args.org, args.cycle_id, args.dry_run)
    logger.info("[%s] Stats: %s", args.org, stats)
    if args.dry_run:
        logger.info("(DRY RUN — nothing was written)")


if __name__ == "__main__":
    main()
