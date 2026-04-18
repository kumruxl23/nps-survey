"""Service for importing nominations from uploaded Excel/CSV files.

Supports .xlsx, .xls, and .csv files. Handles multiple CSV formats:
- Standard format: Name, Email, Leader columns
- Amazon NPS format: PoC, PoC Alias, Stakeholder, Stakeholder Alias,
  "Should this stakeholder be included..." columns.
  Aliases get @amazon.com appended automatically.
"""

import csv
import io
import logging
import re

from app.db import nps_nomination_repo
from app.db.models import ImportResult, Nomination

logger = logging.getLogger(__name__)

# Known column name patterns (case-insensitive)
_EMAIL_PATTERNS = ["email", "email address", "e-mail", "stakeholder email", "stakeholder alias"]
_NAME_PATTERNS = ["name", "full name", "stakeholder name", "stakeholder"]
_LEADER_PATTERNS = ["leader", "poc", "manager"]
_INCLUDE_PATTERNS = ["should this stakeholder be included", "include", "included"]


def import_from_excel(
    org_id: str,
    cycle_id: str,
    file_bytes: bytes,
    filename: str,
) -> ImportResult:
    """Import nominations from an uploaded Excel or CSV file."""
    lower = filename.lower()
    if lower.endswith(".csv"):
        rows = _parse_csv(file_bytes)
    elif lower.endswith((".xlsx", ".xls")):
        rows = _parse_excel(file_bytes)
    else:
        raise ValueError(f"Unsupported file format: {filename}. Use .xlsx, .xls, or .csv")

    total_in_source = len(rows)
    imported_count = 0
    skipped_duplicates = 0
    skipped_excluded = 0

    for row in rows:
        name = row["name"].strip()
        email = row["email"].strip().lower()
        leader = row.get("leader", "").strip()
        include = row.get("include", "yes").strip().lower()

        if not email:
            continue

        # Skip stakeholders marked as "no" for inclusion
        if include in ("no", "n", "false", "0"):
            skipped_excluded += 1
            continue

        # Append @amazon.com if email is just an alias (no @ sign)
        if "@" not in email:
            email = email + "@amazon.com"

        existing = nps_nomination_repo.get_nomination(org_id, cycle_id, email)
        if existing is not None:
            skipped_duplicates += 1
            continue

        nomination = Nomination(
            org_id=org_id,
            cycle_id=cycle_id,
            email=email,
            name=name,
            leader=leader,
        )
        nps_nomination_repo.put_nomination(nomination)
        imported_count += 1

    return ImportResult(
        imported_count=imported_count,
        skipped_duplicates=skipped_duplicates,
        total_in_source=total_in_source,
    )


def _parse_csv(file_bytes: bytes) -> list[dict]:
    """Parse a CSV file into standardized name/email/leader/include dicts.

    Handles duplicate column names by using positional detection.
    """
    text = file_bytes.decode("utf-8-sig")
    lines = text.strip().split("\n")
    if not lines:
        raise ValueError("Empty CSV file")

    # Parse header manually to handle duplicate column names
    reader = csv.reader(io.StringIO(lines[0]))
    raw_headers = next(reader)
    raw_headers = [h.strip() for h in raw_headers]

    # Detect columns by position
    col_map = _detect_columns_positional(raw_headers)

    # Parse data rows
    data_reader = csv.reader(io.StringIO("\n".join(lines[1:])))
    rows = []
    for row_values in data_reader:
        if not row_values or all(not v.strip() for v in row_values):
            continue
        name = row_values[col_map["name_idx"]].strip() if col_map["name_idx"] < len(row_values) else ""
        email = row_values[col_map["email_idx"]].strip() if col_map["email_idx"] < len(row_values) else ""
        leader = row_values[col_map["leader_idx"]].strip() if col_map.get("leader_idx") is not None and col_map["leader_idx"] < len(row_values) else ""
        include = row_values[col_map["include_idx"]].strip() if col_map.get("include_idx") is not None and col_map["include_idx"] < len(row_values) else "yes"
        if email:
            rows.append({"name": name, "email": email, "leader": leader, "include": include})
    return rows


def _parse_excel(file_bytes: bytes) -> list[dict]:
    """Parse an Excel file into standardized name/email/leader/include dicts."""
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active

    header_row = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        header_row.append(str(cell.value or "").strip())

    col_map = _detect_columns(header_row)
    name_idx = header_row.index(col_map["name"])
    email_idx = header_row.index(col_map["email"])
    leader_idx = header_row.index(col_map["leader"]) if col_map.get("leader") and col_map["leader"] in header_row else None
    include_idx = header_row.index(col_map["include"]) if col_map.get("include") and col_map["include"] in header_row else None

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = str(row[name_idx] or "").strip() if name_idx < len(row) else ""
        email = str(row[email_idx] or "").strip() if email_idx < len(row) else ""
        leader = str(row[leader_idx] or "").strip() if leader_idx is not None and leader_idx < len(row) else ""
        include = str(row[include_idx] or "yes").strip() if include_idx is not None and include_idx < len(row) else "yes"
        if email:
            rows.append({"name": name, "email": email, "leader": leader, "include": include})

    wb.close()
    return rows


def _detect_columns(headers: list[str]) -> dict:
    """Auto-detect column mapping from headers. Used by Excel parser."""
    positional = _detect_columns_positional(headers)
    result = {}
    result["name"] = headers[positional["name_idx"]]
    result["email"] = headers[positional["email_idx"]]
    if positional.get("leader_idx") is not None:
        result["leader"] = headers[positional["leader_idx"]]
    if positional.get("include_idx") is not None:
        result["include"] = headers[positional["include_idx"]]
    return result


def _detect_columns_positional(headers: list[str]) -> dict:
    """Auto-detect column indices from headers.

    Handles duplicate column names (e.g., two "Stakeholder" columns)
    by using position-based detection.

    Supports:
    - Standard: Name, Email, Leader
    - Amazon NPS: PoC, PoC Alias, Stakeholder, Stakeholder(alias), Should include...
    """
    lower_headers = [h.lower().strip() for h in headers]
    result = {}

    # Check for Amazon NPS format: PoC, PoC Alias, Stakeholder, Stakeholder(alias)
    # Pattern: if we see "poc" and two "stakeholder" columns, it's the Amazon format
    stakeholder_indices = [i for i, h in enumerate(lower_headers) if "stakeholder" in h and "included" not in h and "respond" not in h and "interact" not in h]

    if len(stakeholder_indices) >= 2:
        # Amazon NPS format: first Stakeholder = name, second Stakeholder = alias/email
        result["name_idx"] = stakeholder_indices[0]
        result["email_idx"] = stakeholder_indices[1]
        logger.info("Detected Amazon NPS format: name=col%d, email=col%d", stakeholder_indices[0], stakeholder_indices[1])
    else:
        # Standard format detection
        # Find email column
        email_idx = None
        for pattern in _EMAIL_PATTERNS:
            for i, h in enumerate(lower_headers):
                if pattern in h:
                    email_idx = i
                    break
            if email_idx is not None:
                break
        if email_idx is None:
            raise ValueError(
                f"Could not find email/alias column. Headers found: {headers}. "
                f"Expected one of: {_EMAIL_PATTERNS} or two 'Stakeholder' columns."
            )
        result["email_idx"] = email_idx

        # Find name column
        name_idx = None
        for pattern in _NAME_PATTERNS:
            for i, h in enumerate(lower_headers):
                if (h == pattern or (pattern in h and "alias" not in h and "included" not in h and "respond" not in h and "interact" not in h)) and i != email_idx:
                    name_idx = i
                    break
            if name_idx is not None:
                break
        if name_idx is None:
            raise ValueError(
                f"Could not find name column. Headers found: {headers}. "
                f"Expected one of: {_NAME_PATTERNS}"
            )
        result["name_idx"] = name_idx

    # Find leader column (PoC)
    for pattern in _LEADER_PATTERNS:
        for i, h in enumerate(lower_headers):
            if (h == pattern or (pattern in h and "alias" not in h)) and i != result.get("name_idx") and i != result.get("email_idx"):
                result["leader_idx"] = i
                break
        if "leader_idx" in result:
            break

    # Find include/exclude column
    for pattern in _INCLUDE_PATTERNS:
        for i, h in enumerate(lower_headers):
            if pattern in h:
                result["include_idx"] = i
                break
        if "include_idx" in result:
            break

    logger.info("Column mapping: %s (from headers: %s)", result, headers)
    return result
