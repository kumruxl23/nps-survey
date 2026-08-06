"""Cycle data export to XLSX (admin download).

Builds a spreadsheet of survey responses for one cycle, with the columns
admins asked for:

    Org | Leader | Stakeholder | NPS Category | What Was Missing | Feedback | Action Taken

Any field that isn't captured is left blank (per requirement) — e.g.
historical anonymous responses have no ``respondent_name``, and
``Action Taken`` maps to the response's internal ``admin_comment`` note,
which may be empty.

Data source: NpsResponse rows (per org + cycle). When ``org_id`` is
omitted, every active org that has responses for the cycle is included.
"""

import io
import logging

from app.services import nps_org_config_service, nps_response_service

logger = logging.getLogger(__name__)

COLUMNS = [
    "Org",
    "Leader",
    "Stakeholder",
    "NPS Category",
    "What Was Missing",
    "Feedback",
    "Action Taken",
]


def _rows_for_org(org, cycle_id: str) -> list[list[str]]:
    rows = []
    for r in nps_response_service.get_responses(org.org_id, cycle_id):
        rows.append([
            org.org_name or org.org_id,
            r.leader or "",
            r.respondent_name or "",
            r.category or "",
            r.what_missing_text or "",
            r.feedback_text or "",
            r.admin_comment or "",
        ])
    return rows


def build_cycle_export(cycle_id: str, org_id: str = "") -> tuple[bytes, str]:
    """Build an XLSX workbook of a cycle's responses.

    Args:
        cycle_id: The cycle to export (required).
        org_id: Restrict to one org; empty = all active orgs.

    Returns:
        (xlsx_bytes, suggested_filename).

    Raises:
        ValueError: when cycle_id is missing.
    """
    cycle_id = (cycle_id or "").strip()
    if not cycle_id:
        raise ValueError("cycle_id is required")

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    org_id = (org_id or "").strip()
    orgs = nps_org_config_service.list_active_orgs()
    if org_id:
        orgs = [o for o in orgs if o.org_id == org_id]

    rows: list[list[str]] = []
    for org in orgs:
        rows.extend(_rows_for_org(org, cycle_id))

    wb = Workbook()
    ws = wb.active
    ws.title = "NPS Responses"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    ws.append(COLUMNS)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    for row in rows:
        ws.append(row)

    # Reasonable column widths; wrap the long free-text columns.
    widths = [16, 22, 24, 14, 40, 48, 40]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = width
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    scope = org_id or "all-orgs"
    filename = f"nps_export_{scope}_{cycle_id}.xlsx"
    logger.info("Built cycle export: cycle=%s scope=%s rows=%d", cycle_id, scope, len(rows))
    return buf.getvalue(), filename
