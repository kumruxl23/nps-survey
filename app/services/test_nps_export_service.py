"""Tests for the cycle XLSX export service."""

import io
from types import SimpleNamespace
from unittest.mock import patch

from openpyxl import load_workbook

from app.services import nps_export_service
from app.db.models import NpsResponse


def _resp(**kw):
    base = dict(
        org_id="whs_cpt_in", cycle_id="h2-2026", response_id="r1",
        nps_score=9, category="Promoter", leader="Navjyot Bhatia",
        feedback_text="great support", what_missing_text="",
        respondent_name="Asha K", admin_comment="scheduled a sync",
    )
    base.update(kw)
    return NpsResponse(**base)


def _read(data: bytes):
    wb = load_workbook(io.BytesIO(data))
    ws = wb.active
    # openpyxl reads empty cells as None; normalize to "" (blank in Excel).
    return [[(c.value if c.value is not None else "") for c in row] for row in ws.iter_rows()]


def test_requires_cycle_id():
    import pytest
    with pytest.raises(ValueError, match="cycle_id"):
        nps_export_service.build_cycle_export("")


@patch("app.services.nps_export_service.nps_response_service.get_responses")
@patch("app.services.nps_export_service.nps_org_config_service.list_active_orgs")
def test_headers_and_rows(mock_orgs, mock_get):
    mock_orgs.return_value = [SimpleNamespace(org_id="whs_cpt_in", org_name="WHS CPT IN")]
    mock_get.return_value = [_resp()]

    data, filename = nps_export_service.build_cycle_export("h2-2026")
    rows = _read(data)

    assert rows[0] == nps_export_service.COLUMNS
    assert rows[1] == ["WHS CPT IN", "Navjyot Bhatia", "Asha K", "Promoter",
                       "", "great support", "scheduled a sync"]
    assert filename == "nps_export_all-orgs_h2-2026.xlsx"


@patch("app.services.nps_export_service.nps_response_service.get_responses")
@patch("app.services.nps_export_service.nps_org_config_service.list_active_orgs")
def test_blank_fields_left_empty(mock_orgs, mock_get):
    mock_orgs.return_value = [SimpleNamespace(org_id="fec", org_name="FEC")]
    # anonymous response: no name, no feedback, no action
    mock_get.return_value = [_resp(respondent_name="", feedback_text="",
                                   what_missing_text="", admin_comment="", leader="")]
    data, _ = nps_export_service.build_cycle_export("h2-2026", org_id="fec")
    rows = _read(data)
    assert rows[1] == ["FEC", "", "", "Promoter", "", "", ""]


@patch("app.services.nps_export_service.nps_response_service.get_responses")
@patch("app.services.nps_export_service.nps_org_config_service.list_active_orgs")
def test_org_filter_scopes_filename(mock_orgs, mock_get):
    mock_orgs.return_value = [
        SimpleNamespace(org_id="whs_cpt_in", org_name="WHS CPT IN"),
        SimpleNamespace(org_id="fec", org_name="FEC"),
    ]
    mock_get.return_value = []
    data, filename = nps_export_service.build_cycle_export("h2-2026", org_id="fec")
    # only the fec org should be queried
    assert mock_get.call_count == 1
    assert filename == "nps_export_fec_h2-2026.xlsx"
