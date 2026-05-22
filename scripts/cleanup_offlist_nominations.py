"""Clean up two specific off-list-nomination patterns.

Pattern 1: Synthetic-email nominations whose stakeholder IS in the
           workbook Stakeholder List by name.
  Action: rewrite the nomination using the real email from the
          Stakeholder List. The leader and responded flag are preserved.
  Why: the original backfill couldn't match them when their Asana task
       had no assignee email set, so it fabricated
       `asana-task-{gid}@unknown.local`. We now know who they are by
       name; map back to their real address so reminders + dedup work.

Pattern 2: Leader self-responses — nominations whose name is the same
           as the leader they're tagged against.
  Action: delete the nomination row. Leave the corresponding NpsResponse
          rows untouched (their feedback is still valuable).

Usage on the EC2:

    /usr/bin/python3.11 scripts/cleanup_offlist_nominations.py \\
        --workbook /tmp/cpt_na.xlsx --org whs_cpt_na --cycle h1-2026 --dry-run

Drop --dry-run to apply. Run for each (workbook, org) you want to clean.

Safety:
  - --dry-run prints planned actions, writes nothing.
  - Aborts before any mutation if the workbook parses 0 stakeholder names.
  - Never touches NpsResponses.
"""

from __future__ import annotations

import argparse
import logging
import os
import sys
from pathlib import Path

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
os.environ.setdefault("AWS_REGION", "ap-south-1")
os.environ.setdefault("AWS_DEFAULT_REGION", "ap-south-1")

from app.db import nps_nomination_repo  # noqa: E402
from app.db.models import Nomination  # noqa: E402
from app.services.nomination_keys import base_email, encode_for_leader  # noqa: E402
from scripts.import_h1_2026_stakeholders import (  # noqa: E402
    ORG_CONFIGS,
    _norm_leader_name,
    _parse_stakeholder_list,
)

logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
logger = logging.getLogger("cleanup")

SYNTHETIC_EMAIL_PREFIX = "asana-task-"
SYNTHETIC_EMAIL_DOMAIN = "@unknown.local"


def _build_name_to_email(workbook: Path, org: str) -> dict[str, str]:
    """Map normalized stakeholder name -> email from the workbook's
    Stakeholder List sheet. Names are lowercased + whitespace-collapsed."""
    cfg = ORG_CONFIGS[org]
    wb = sys.modules["openpyxl"].load_workbook(workbook, data_only=True)  # noqa: F401
    sh_map = _parse_stakeholder_list(workbook, cfg)
    # sh_map is keyed by alias; we need to also index by the stakeholder
    # NAME so we can match Asana's "John Sarreal" -> hallberg@amazon.com.
    # Re-walk the sheet manually to get name-keyed entries too.
    import openpyxl
    wb = openpyxl.load_workbook(workbook, data_only=True)
    sheet_name = cfg["stakeholder_list_sheet_name"]
    if sheet_name not in wb.sheetnames:
        return {}
    ws = wb[sheet_name]
    pos = cfg["stakeholder_list_positions"]
    leader_pos = pos["leader_pos"]
    alias_pos = pos["stakeholder_alias_pos"]
    email_pos = pos["email_pos"]

    # Stakeholder Name is usually at position 2 (CPT NA, FEC) or
    # mixed in CPT IN; figure out by scanning headers.
    # Heuristic: stakeholder name column is the one with full names
    # (alphabetic strings with spaces), AT a position other than the
    # alias position. Simplest: try positions 2, then 3, then 1.
    name_to_email: dict[str, str] = {}

    header_row = None
    for r in range(1, min(ws.max_row, 10) + 1):
        nonempty = sum(1 for c in ws[r] if c.value not in (None, ""))
        if nonempty >= 3:
            header_row = r
            break
    if header_row is None:
        return {}

    # Heuristic to find name col: the column we use for alias is usually 3,
    # email is 4, leader is 0, poc is 1, and the stakeholder NAME column
    # is wherever has full names — typically 2 for FEC/CPT-NA.
    # We'll just search columns 0..(email_pos-1) for the one with the
    # most space-containing text values.
    candidate_cols = [c for c in range(email_pos) if c not in (leader_pos, alias_pos)]
    best_col = None
    best_score = -1
    for c in candidate_cols:
        score = 0
        for r in range(header_row + 1, min(ws.max_row, header_row + 30) + 1):
            v = ws.cell(r, c + 1).value
            if isinstance(v, str) and " " in v.strip():
                score += 1
        if score > best_score:
            best_col = c
            best_score = score

    if best_col is None:
        return {}

    for r in range(header_row + 1, ws.max_row + 1):
        values = [c.value for c in ws[r]]
        if not any(v not in (None, "") for v in values):
            continue
        raw_name = values[best_col] if best_col < len(values) else None
        email_raw = values[email_pos] if email_pos < len(values) else None
        if not raw_name:
            continue
        # Strip post-comma title (e.g. "Alex Tribendis, Director WHS" -> "Alex Tribendis")
        name = str(raw_name).split(",", 1)[0].strip()
        name_key = " ".join(name.lower().split())
        email = (str(email_raw) if email_raw else "").strip().lower()
        if email and name_key not in name_to_email:
            name_to_email[name_key] = email

    return name_to_email


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", type=Path, required=True)
    parser.add_argument("--org", required=True, choices=sorted(ORG_CONFIGS.keys()))
    parser.add_argument("--cycle", required=True)
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    if not args.workbook.is_file():
        sys.exit(f"Workbook not found: {args.workbook}")

    name_to_email = _build_name_to_email(args.workbook, args.org)
    if not name_to_email:
        sys.exit(f"Workbook returned 0 name->email entries — refusing to mutate. "
                 f"Check workbook + org='{args.org}' setup.")
    logger.info("[%s] %d name->email entries from Stakeholder List", args.org, len(name_to_email))

    nominations = nps_nomination_repo.list_nominations(args.org, args.cycle)
    logger.info("[%s] %d nominations in DynamoDB", args.org, len(nominations))

    pattern1_actions = []  # (old_nom, new_email)
    pattern2_actions = []  # nominations to delete (leader self-noms)

    # Build a set of currently-taken keys so the rewrite can pick a unique key
    taken_keys = {n.email.strip().lower() for n in nominations if n.email}

    for n in nominations:
        # Pattern 2 — leader self-response: nomination.name matches nomination.leader
        nom_name = _norm_leader_name((n.name or "").strip())
        nom_leader = _norm_leader_name((n.leader or "").strip())
        if nom_name and nom_leader and nom_name == nom_leader:
            pattern2_actions.append(n)
            continue

        # Pattern 1 — synthetic email AND name in workbook
        if n.email and n.email.startswith(SYNTHETIC_EMAIL_PREFIX) and SYNTHETIC_EMAIL_DOMAIN in n.email:
            name_key = " ".join((n.name or "").lower().split())
            real_email = name_to_email.get(name_key)
            if real_email:
                pattern1_actions.append((n, real_email))

    logger.info("[%s] Pattern 1 (synthetic -> real email): %d candidates",
                args.org, len(pattern1_actions))
    for old, new_email in pattern1_actions:
        logger.info("  rewrite: %s -> %s (name='%s' leader='%s')",
                    old.email, new_email, old.name, old.leader)

    logger.info("[%s] Pattern 2 (leader self-response, drop nomination): %d candidates",
                args.org, len(pattern2_actions))
    for n in pattern2_actions:
        logger.info("  delete: email=%s name='%s' leader='%s'", n.email, n.name, n.leader)

    if args.dry_run:
        logger.info("DRY RUN — no changes made.")
        return

    # Apply Pattern 1: delete old synthetic row, write new with real email.
    # Use encode_for_leader so a multi-leader collision still produces a
    # unique key.
    for old, new_email in pattern1_actions:
        # Remove the old key from the taken set since we're deleting it
        taken_keys.discard(old.email.strip().lower())
        new_key = encode_for_leader(new_email, old.leader, taken_keys)
        new_nom = Nomination(
            org_id=old.org_id,
            cycle_id=old.cycle_id,
            email=new_key,
            name=old.name,
            leader=old.leader,
            slack_user_id=old.slack_user_id,
            responded=old.responded,
            responded_at=old.responded_at,
            created_at=old.created_at,
        )
        nps_nomination_repo.delete_nomination(args.org, args.cycle, old.email)
        nps_nomination_repo.put_nomination(new_nom)

    # Apply Pattern 2: just delete the nomination
    for n in pattern2_actions:
        nps_nomination_repo.delete_nomination(args.org, args.cycle, n.email)

    logger.info("[%s] Done. rewrote=%d deleted=%d",
                args.org, len(pattern1_actions), len(pattern2_actions))


if __name__ == "__main__":
    main()
